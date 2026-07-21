"""
Serviço CNAE: carrega a tabela proprietária (NOVO CNAE.xlsx) em memória no startup
e fornece enriquecimento de códigos CNAE com hierarquia Setor→Segmento→Ramo→Categoria→Produto.
"""
from __future__ import annotations

import os
from pathlib import Path
from typing import Dict, Optional

from loguru import logger

# Caminho padrão — pode ser sobrescrito via env CNAE_XLSX_PATH
_DEFAULT_PATH = Path(__file__).parent.parent / "data" / "NOVO_CNAE.xlsx"

# Tabela BigQuery com a hierarquia CNAE (fonte primária; evita ler Excel na startup)
_DEFAULT_BQ_TABLE = "liquid-receiver-483923-n6.Projeto_Comex.cnae_hierarquia"

# Dicionário global carregado no startup
_cnae_map: Dict[str, dict] = {}
_loaded = False


def _normalizar_cnae(codigo) -> str:
    """Converte CNAE para string de 7 dígitos sem formatação."""
    return str(codigo).strip().replace(".", "").replace("-", "").replace("/", "").lstrip("0").zfill(7)


def carregar_cnae(path: Optional[str] = None) -> int:
    """
    Lê o arquivo XLSX e popula o mapa global.
    Retorna o número de códigos carregados.
    Seguro para chamar múltiplas vezes (idempotente).
    """
    global _cnae_map, _loaded

    # 1) Fonte primária: BigQuery (leve, sem pandas/openpyxl). Só é ignorada se
    #    o caller passar um path explícito de Excel.
    if path is None and os.getenv("CNAE_USE_BQ", "true").strip().lower() in {"1", "true", "yes", "y"}:
        try:
            from services.bq_client import get_bigquery_client, run_query
            table = os.getenv("CNAE_BQ_TABLE", _DEFAULT_BQ_TABLE).strip().strip("`")
            client = get_bigquery_client()
            rows = run_query(client, f"SELECT cnae, descricao, setor, segmento, ramo, categoria, produto FROM `{table}`", None)
            mapa: Dict[str, dict] = {}
            for r in rows:
                codigo = _normalizar_cnae(r.get("cnae") or "")
                if not codigo or codigo == "0000000":
                    continue
                mapa[codigo] = {
                    "descricao": r.get("descricao"),
                    "setor": r.get("setor"),
                    "segmento": r.get("segmento"),
                    "ramo": r.get("ramo"),
                    "categoria": r.get("categoria"),
                    "produto": r.get("produto"),
                }
            if mapa:
                _cnae_map = mapa
                _loaded = True
                logger.info(f"✅ CNAE carregado do BigQuery: {len(mapa)} códigos ({table})")
                return len(mapa)
            logger.warning("CNAE BigQuery retornou vazio — caindo para o Excel.")
        except Exception as exc:
            logger.warning(f"CNAE BigQuery indisponível ({exc}); usando Excel local.")

    xlsx_path = Path(path or os.getenv("CNAE_XLSX_PATH") or _DEFAULT_PATH)
    if not xlsx_path.exists():
        logger.warning(f"⚠️  Arquivo CNAE não encontrado em {xlsx_path} — enriquecimento desabilitado.")
        _loaded = True
        return 0

    try:
        # openpyxl read-only em vez de pandas: evita carregar pandas+numpy na
        # startup (economiza ~60MB de RAM no Render free de 512MB).
        from openpyxl import load_workbook

        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header = [str(h).strip() if h is not None else "" for h in next(it)]
        idx = {name: i for i, name in enumerate(header)}

        def _get(row, col):
            i = idx.get(col)
            return row[i] if (i is not None and i < len(row)) else None

        def _clean(v):
            s = str(v).strip() if v is not None else ""
            return s if s and s.lower() not in ("nan", "none") else None

        mapa: Dict[str, dict] = {}
        for row in it:
            codigo = _normalizar_cnae(_get(row, "CNAE") or "")
            if not codigo or codigo == "0000000":
                continue
            mapa[codigo] = {
                "descricao": _clean(_get(row, "DESCRIÇÃO")),
                "setor": _clean(_get(row, "SETOR")),
                "segmento": _clean(_get(row, "SEGMENTO")),
                "ramo": _clean(_get(row, "RAMO")),
                "categoria": _clean(_get(row, "CATEGORIA")),
                "produto": _clean(_get(row, "PRODUTO")),
            }
        wb.close()

        _cnae_map = mapa
        _loaded = True
        logger.info(f"✅ CNAE carregado: {len(mapa)} códigos de {xlsx_path.name}")
        return len(mapa)

    except Exception as exc:
        logger.error(f"❌ Erro ao carregar CNAE: {exc}")
        _loaded = True
        return 0


def enriquecer(cnae_codigo) -> Optional[dict]:
    """
    Retorna o dict de hierarquia para o código CNAE ou None se não encontrado.
    Tenta variações com zeros à esquerda (6 e 7 dígitos).
    """
    if not _loaded:
        carregar_cnae()

    if not cnae_codigo:
        return None

    # Tentar normalização padrão (7 dígitos)
    chave7 = _normalizar_cnae(cnae_codigo)
    if chave7 in _cnae_map:
        return _cnae_map[chave7]

    # Tentar 6 dígitos (alguns CNAEs da RF têm 6 dígitos sem o dígito verificador)
    chave6 = chave7.lstrip("0").zfill(6)
    for k in (chave6, chave6.lstrip("0")):
        if k in _cnae_map:
            return _cnae_map[k]

    # Busca prefixo (primeiros 4 dígitos) como fallback
    prefixo = chave7[:4]
    for k, v in _cnae_map.items():
        if k.startswith(prefixo):
            return v

    return None


def listar_setores() -> list[str]:
    return sorted({v["setor"] for v in _cnae_map.values() if v.get("setor")})


def listar_segmentos(setor: Optional[str] = None) -> list[str]:
    return sorted({
        v["segmento"] for v in _cnae_map.values()
        if v.get("segmento") and (not setor or v.get("setor") == setor)
    })


def enriquecer_por_prefixo(cnae_codigo, setor: Optional[str] = None, segmento: Optional[str] = None,
                           ramo: Optional[str] = None, categoria: Optional[str] = None) -> Optional[dict]:
    """Enriquecimento por prefixo CNAE de 4 dígitos (classe). Robusto a formatos
    como '07.21-9'. Quando há filtro ativo (setor/segmento/...), prefere a
    hierarquia que casa com ele (um prefixo pode mapear a múltiplos segmentos)."""
    if not _loaded:
        carregar_cnae()
    digitos = "".join(c for c in str(cnae_codigo or "") if c.isdigit())
    if len(digitos) < 4:
        return None
    pref = digitos[:4]
    candidatos = [
        v for k, v in _cnae_map.items()
        if "".join(c for c in str(k) if c.isdigit())[:4] == pref
    ]
    if not candidatos:
        return None
    for v in candidatos:
        if ((not setor or v.get("setor") == setor)
                and (not segmento or v.get("segmento") == segmento)
                and (not ramo or v.get("ramo") == ramo)
                and (not categoria or v.get("categoria") == categoria)):
            return v
    return candidatos[0]


def listar_ramos(setor: Optional[str] = None, segmento: Optional[str] = None) -> list:
    if not _loaded:
        carregar_cnae()
    return sorted({
        v["ramo"] for v in _cnae_map.values()
        if v.get("ramo")
        and (not setor or v.get("setor") == setor)
        and (not segmento or v.get("segmento") == segmento)
    })


def listar_categorias(setor: Optional[str] = None, segmento: Optional[str] = None,
                      ramo: Optional[str] = None) -> list:
    if not _loaded:
        carregar_cnae()
    return sorted({
        v["categoria"] for v in _cnae_map.values()
        if v.get("categoria")
        and (not setor or v.get("setor") == setor)
        and (not segmento or v.get("segmento") == segmento)
        and (not ramo or v.get("ramo") == ramo)
    })


def arvore() -> Dict[str, dict]:
    """Árvore aninhada Setor → Segmento → Ramo → [Categorias] para dropdowns."""
    if not _loaded:
        carregar_cnae()
    tree: Dict[str, dict] = {}
    for v in _cnae_map.values():
        s = v.get("setor")
        if not s:
            continue
        seg = v.get("segmento") or "—"
        ramo = v.get("ramo") or "—"
        cat = v.get("categoria")
        tree.setdefault(s, {}).setdefault(seg, {}).setdefault(ramo, set())
        if cat:
            tree[s][seg][ramo].add(cat)
    # set → lista ordenada
    return {
        s: {seg: {r: sorted(cats) for r, cats in ramos.items()} for seg, ramos in segs.items()}
        for s, segs in tree.items()
    }


def prefixos_por_filtro(setor: Optional[str] = None, segmento: Optional[str] = None,
                        ramo: Optional[str] = None, categoria: Optional[str] = None) -> list:
    """Retorna os prefixos CNAE de 4 dígitos que correspondem ao filtro de hierarquia.
    Usado para casar com a base de empresas (CNAE a nível de classe)."""
    if not _loaded:
        carregar_cnae()
    prefixos = set()
    for codigo, v in _cnae_map.items():
        if setor and v.get("setor") != setor:
            continue
        if segmento and v.get("segmento") != segmento:
            continue
        if ramo and v.get("ramo") != ramo:
            continue
        if categoria and v.get("categoria") != categoria:
            continue
        digitos = "".join(c for c in str(codigo) if c.isdigit())
        if len(digitos) >= 4:
            prefixos.add(digitos[:4])
    return sorted(prefixos)


def cnae_map() -> Dict[str, dict]:
    if not _loaded:
        carregar_cnae()
    return _cnae_map
